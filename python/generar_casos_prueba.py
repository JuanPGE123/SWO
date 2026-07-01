"""
Generador de casos de prueba GA9-220501096-AA2-EV01
Proyecto SWO (ServiceDesk) — Formato GFPI-F-135 V01
"""
from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter

# ─── Paleta de colores ───────────────────────────────────────────────────────
COLOR_HEADER_BG    = "0B1116"   # azul marino oscuro (navbar SWO)
COLOR_HEADER_FG    = "05D0E6"   # cian SWO
COLOR_CAMPO_BG     = "1A2632"   # azul oscuro secundario
COLOR_CAMPO_FG     = "FFFFFF"   # blanco
COLOR_VALOR_BG     = "F0F4F8"   # gris muy claro
COLOR_APROBADO     = "1E7E34"   # verde oscuro
COLOR_SEGUIMIENTO  = "856404"   # ámbar oscuro
COLOR_RECHAZADO    = "721C24"   # rojo oscuro
COLOR_ALTO         = "721C24"
COLOR_MEDIO        = "856404"
COLOR_BAJO         = "155724"
COLOR_RESUMEN_HDR  = "0B1116"
COLOR_RESUMEN_FG   = "05D0E6"
COLOR_ALT_ROW      = "E8F4F8"   # fila alternada resumen

THIN = Side(style="thin", color="AAAAAA")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
THICK_BOTTOM = Border(
    left=THIN, right=THIN,
    top=THIN, bottom=Side(style="medium", color="05D0E6")
)


def cell_style(ws, row, col, value,
               bold=False, bg=None, fg="000000",
               wrap=True, align_h="left", align_v="center",
               font_size=10, border=True):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(bold=bold, color=fg, size=font_size,
                  name="Calibri")
    c.alignment = Alignment(
        horizontal=align_h, vertical=align_v,
        wrap_text=wrap
    )
    if bg:
        c.fill = PatternFill("solid", fgColor=bg)
    if border:
        c.border = BORDER
    return c


def write_case(wb, case_number, case_data):
    title = f"CP-{case_number:03d}"
    ws = wb.create_sheet(title=title)

    # Ancho de columnas
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 90

    # ── Fila 1: Encabezado del documento ────────────────────────────────────
    ws.merge_cells("A1:B1")
    c = ws.cell(row=1, column=1,
                value="FORMATO DE CASO DE PRUEBA — GFPI-F-135 V01")
    c.font = Font(bold=True, color=COLOR_HEADER_FG, size=13, name="Calibri")
    c.fill = PatternFill("solid", fgColor=COLOR_HEADER_BG)
    c.alignment = Alignment(horizontal="center", vertical="center",
                            wrap_text=False)
    c.border = THICK_BOTTOM
    ws.row_dimensions[1].height = 30

    # ── Fila 2: Subtítulo proyecto ───────────────────────────────────────────
    ws.merge_cells("A2:B2")
    c = ws.cell(row=2, column=1,
                value="SWO (ServiceDesk)  ·  GA9-220501096-AA2-EV01")
    c.font = Font(bold=False, color="CCCCCC", size=9, name="Calibri")
    c.fill = PatternFill("solid", fgColor=COLOR_HEADER_BG)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = BORDER
    ws.row_dimensions[2].height = 18

    # ── Filas de datos ───────────────────────────────────────────────────────
    start_row = 3
    for i, (campo, valor) in enumerate(case_data):
        row = start_row + i
        ws.row_dimensions[row].height = 50

        # Columna A — nombre del campo
        cell_style(ws, row, 1, campo,
                   bold=True, bg=COLOR_CAMPO_BG, fg=COLOR_HEADER_FG,
                   align_h="left", font_size=10)

        # Columna B — valor
        bg_val = COLOR_VALOR_BG if i % 2 == 0 else "FFFFFF"

        # Colorear celdas especiales
        if campo == "Resultado":
            if "Rechazado" in str(valor):
                bg_val = "F8D7DA"
                fg_val = COLOR_RECHAZADO
            elif "seguimiento" in str(valor).lower():
                bg_val = "FFF3CD"
                fg_val = COLOR_SEGUIMIENTO
            else:
                bg_val = "D4EDDA"
                fg_val = COLOR_APROBADO
            cell_style(ws, row, 2, valor,
                       bold=True, bg=bg_val, fg=fg_val,
                       align_h="center", font_size=11)
        elif campo == "Severidad":
            if "Alto" in str(valor):
                bg_val, fg_val = "F8D7DA", COLOR_ALTO
            elif "Medio" in str(valor):
                bg_val, fg_val = "FFF3CD", COLOR_SEGUIMIENTO
            else:
                bg_val, fg_val = "D4EDDA", COLOR_BAJO
            cell_style(ws, row, 2, valor,
                       bold=True, bg=bg_val, fg=fg_val,
                       align_h="center", font_size=10)
        elif campo == "Número de caso":
            cell_style(ws, row, 2, valor,
                       bold=True, bg=bg_val, fg="0B1116",
                       align_h="center", font_size=12)
        else:
            cell_style(ws, row, 2, valor,
                       bg=bg_val, fg="0B1116",
                       align_h="left", font_size=10)

    # Ajustar altura de filas con mucho texto
    for row in ws.iter_rows():
        for cell in row:
            if cell.value and len(str(cell.value)) > 150:
                ws.row_dimensions[cell.row].height = 120
            elif cell.value and len(str(cell.value)) > 80:
                ws.row_dimensions[cell.row].height = 80


# ─── Datos de los 5 casos ────────────────────────────────────────────────────

CASOS = [
    # ── CASO 001 ──────────────────────────────────────────────────────────────
    [
        ("Nombre del proyecto",       "SWO (ServiceDesk)"),
        ("Nombre del caso de prueba", "Resolver incidente con notas técnicas obligatorias en modo offline — App Android"),
        ("Fecha de revisión",         "11/06/2026"),
        ("Descripción del caso",
         "Prueba de resiliencia offline (Offline-First). Se ejecuta sobre la app nativa Android de SWO. "
         "El técnico marca un incidente como 'Resuelto' y diligencia notas técnicas de cierre mientras el "
         "dispositivo está en Modo Avión. Se verifica que Room persista el registro y que WorkManager encole "
         "la sincronización, ejecutándola al recuperar conectividad. Fin: garantizar continuidad operativa "
         "en sitios sin cobertura de red."),
        ("Ambiente o entorno de prueba",
         "SO Host: Windows 11 Pro 22H2 | Emulador: Android Studio (Pixel 7 API 34 — Android 14) | "
         "Conectividad: deshabilitada con Modo Avión | Backend Spring Boot en Railway o localhost:8080 | "
         "BD local: Room v2.6 + SQLite | WorkManager v2.9 para sincronización diferida."),
        ("Herramienta utilizada",
         "Android Emulator (Android Studio Hedgehog 2023.1.1) · Espresso UI Test Framework · "
         "ADB para captura de logs · Logcat (filtro tag: WorkManager, SWO-Sync)"),
        ("Autor del caso de prueba",  "Juan Pablo Giraldo Echavarria"),
        ("Número de caso",            "001"),
        ("Salida esperada",
         "1) Incidente marcado como 'Resuelto' localmente en Room sin errores.\n"
         "2) Indicador visual 'Pendiente de sincronización' visible en UI.\n"
         "3) Al restaurar red, WorkManager ejecuta el job; backend responde HTTP 200 PUT /v1/incidencias/{id}.\n"
         "4) Indicador de sincronización desaparece y el historial muestra fecha/hora de cierre correcta."),
        ("Salida obtenida",
         "Incidente guardado localmente en Room sin errores (verificado en Logcat). Al restaurar la red, "
         "WorkManager presenta retraso de ~45 s antes de sincronizar. El backend responde HTTP 200 OK. "
         "Sin embargo, la timestamp de cierre muestra la hora de reconexión, no la hora real offline de la acción."),
        ("Resultado",                 "En seguimiento"),
        ("Seguimiento",               "BUG-MOV-001: Timestamp de resolución offline registra hora de sincronización en lugar de hora local de la acción."),
        ("Severidad",                 "Medio"),
        ("Evidencia",
         "Captura de Logcat con tag SWO-Sync mostrando delta de tiempo entre resolución local y sincronización remota. "
         "Screenshot de UI del incidente con discrepancia en fecha/hora de cierre."),
        ("Firma de aprobación",       "Juan Pablo Giraldo Echavarria (Analista de QA)"),
    ],

    # ── CASO 002 ──────────────────────────────────────────────────────────────
    [
        ("Nombre del proyecto",       "SWO (ServiceDesk)"),
        ("Nombre del caso de prueba", "Validación de campos obligatorios al crear incidente en el formulario web Angular 17"),
        ("Fecha de revisión",         "11/06/2026"),
        ("Descripción del caso",
         "Prueba funcional de validación de frontend. Se ejecuta con Cypress sobre IncidentFormComponent. "
         "Se intenta enviar el formulario reactivo de 'Crear Incidente' dejando en blanco los campos requeridos: "
         "Título (Validators.required, noWhitespace, mín. 5 chars), Descripción (mín. 10 palabras, máx. 500 chars) "
         "y Categoría (Validators.required). Fin: verificar que onFormSubmit() active 'touched' y getErrorMessage() "
         "retorne mensajes correctos sin invocar el backend."),
        ("Ambiente o entorno de prueba",
         "SO: Windows 11 Pro 22H2 | Navegadores: Chrome 124 y Edge 124 | Angular 17.3 con ReactiveFormsModule | "
         "Servidor: ng serve en localhost:4200 | Backend: no requerido (interceptor HTTP mockeado) | Node.js 20 LTS"),
        ("Herramienta utilizada",
         "Cypress 13.x (E2E) · Angular DevTools (árbol de formularios reactivos) · "
         "Chrome DevTools — Console para verificar ausencia de peticiones HTTP a /v1/incidencias"),
        ("Autor del caso de prueba",  "Juan Pablo Giraldo Echavarria"),
        ("Número de caso",            "002"),
        ("Salida esperada",
         "1) Sin envío HTTP al backend al hacer clic en 'Guardar' con campos vacíos.\n"
         "2) Título muestra: 'Este campo es obligatorio'.\n"
         "3) Descripción muestra: 'Debe contener al menos 10 palabras'.\n"
         "4) Categoría muestra: 'Este campo es obligatorio'.\n"
         "5) Campos con error presentan borde rojo.\n"
         "6) Título con solo espacios muestra: 'No puede estar vacío o contener solo espacios'."),
        ("Salida obtenida",
         "Título y Categoría muestran sus errores correctamente. Sin embargo, Descripción no activa el validador "
         "minWords cuando el campo está vacío; solo muestra el error 'required', omitiendo el mensaje de mínimo de palabras. "
         "No se realizan peticiones HTTP (verificado en Network tab)."),
        ("Resultado",                 "En seguimiento"),
        ("Seguimiento",               "BUG-WEB-002: CustomValidators.minWords(10) no se evalúa con campo vacío; el error 'required' absorbe la validación y el mensaje de mínimo de palabras nunca es visible."),
        ("Severidad",                 "Bajo"),
        ("Evidencia",
         "Captura del formulario con tres campos en estado de error. Screenshot de Network tab con 0 peticiones "
         "a /v1/incidencias. Reporte HTML de Cypress con assertion fallida en el mensaje de mínimo de palabras."),
        ("Firma de aprobación",       "Juan Pablo Giraldo Echavarria (Analista de QA)"),
    ],

    # ── CASO 003 ──────────────────────────────────────────────────────────────
    [
        ("Nombre del proyecto",       "SWO (ServiceDesk)"),
        ("Nombre del caso de prueba", "Control de acceso por roles — Intento de eliminación de proyecto con perfil 'Cliente'"),
        ("Fecha de revisión",         "11/06/2026"),
        ("Descripción del caso",
         "Prueba de seguridad de autorización (RBAC). Se evalúan dos capas: frontend (Cypress) y API (Postman). "
         "Un usuario con rol 'Cliente' intenta ejecutar DELETE /v1/proyectos/{id}. Fin: verificar que Spring Security "
         "rechace la operación con HTTP 403 y que el frontend Angular no muestre el control de eliminación a roles sin privilegios."),
        ("Ambiente o entorno de prueba",
         "SO: Windows 11 Pro 22H2 | Navegador: Chrome 124 | Backend Spring Boot 3.x en Railway o localhost:8080 | "
         "Spring Security con SecurityConfig.java | BD: PostgreSQL en Railway | "
         "Usuario de prueba con rol 'Cliente' preconfigurado en DataInitializer.java"),
        ("Herramienta utilizada",
         "Postman v11 (colección SWO API con variable {{token_cliente}}) · "
         "Cypress 13.x (visibilidad de controles UI por rol) · "
         "Swagger UI en /swagger-ui.html para documentar endpoint auditado"),
        ("Autor del caso de prueba",  "Juan Pablo Giraldo Echavarria"),
        ("Número de caso",            "003"),
        ("Salida esperada",
         "1) Frontend NO renderiza el botón 'Eliminar Proyecto' para rol 'Cliente'.\n"
         "2) Acceso directo a ruta de administración redirige a /dashboard o muestra 'Acceso denegado'.\n"
         "3) DELETE /v1/proyectos/{id} con credenciales de 'Cliente' retorna HTTP 403 Forbidden con mensaje de error."),
        ("Salida obtenida",
         "HALLAZGO CRÍTICO: El backend retorna HTTP 200 OK y elimina el proyecto con token de rol 'Cliente'. "
         "SecurityConfig.java línea 43 tiene anyRequest().permitAll() con TODO de JWT pendiente. "
         "El authGuard solo verifica autenticación, no rol. El frontend oculta el botón, pero la API no tiene protección."),
        ("Resultado",                 "Rechazado"),
        ("Seguimiento",               "BUG-SEC-003 (CRÍTICO): DELETE /v1/proyectos/{id} no implementa autorización por rol. Cualquier usuario autenticado puede eliminar proyectos. Causa raíz: SecurityConfig.java:43 — anyRequest().permitAll(). Requiere JWT + @PreAuthorize('hasRole(ADMIN)') en ProyectoController."),
        ("Severidad",                 "Alto"),
        ("Evidencia",
         "Screenshot de Postman con DELETE y token de rol 'Cliente' obteniendo HTTP 200 OK con confirmación de eliminación. "
         "Log de Spring Boot sin error de autorización. Fragmento de SecurityConfig.java línea 43 resaltado."),
        ("Firma de aprobación",       "Juan Pablo Giraldo Echavarria (Analista de QA)"),
    ],

    # ── CASO 004 ──────────────────────────────────────────────────────────────
    [
        ("Nombre del proyecto",       "SWO (ServiceDesk)"),
        ("Nombre del caso de prueba", "Manejo de mensajes simultáneos en el Chatbot y estado visual 'El bot está escribiendo…'"),
        ("Fecha de revisión",         "11/06/2026"),
        ("Descripción del caso",
         "Prueba funcional de concurrencia y estado asíncrono. Se ejecuta con Cypress sobre ChatbotComponent. "
         "El evaluador envía múltiples mensajes rápidamente mientras el bot procesa (escribiendo=true). "
         "Se verifica que el guardián 'if (!texto || this.escribiendo) return' bloquee envíos duplicados y que "
         "el BehaviorSubject escribiendo$ maneje correctamente las transiciones de estado. "
         "Fin: garantizar que no haya condiciones de carrera ni mensajes duplicados."),
        ("Ambiente o entorno de prueba",
         "SO: Windows 11 Pro 22H2 | Navegadores: Chrome 124 y Edge 124 | Angular 17 en localhost:4200 | "
         "Backend Spring Boot con ChatbotController en localhost:8080 o Railway | "
         "Latencia simulada: Chrome DevTools throttling a 'Slow 3G' (~400ms) para provocar ventana donde escribiendo=true"),
        ("Herramienta utilizada",
         "Cypress 13.x (comandos .type() y .click() encadenados en rápida sucesión) · "
         "Chrome DevTools — Network throttling (Slow 3G) · "
         "Angular DevTools — inspección del BehaviorSubject escribiendo$ en tiempo real"),
        ("Autor del caso de prueba",  "Juan Pablo Giraldo Echavarria"),
        ("Número de caso",            "004"),
        ("Salida esperada",
         "1) Al enviar el primer mensaje, aparece 'El bot está escribiendo...' (escribiendo=true).\n"
         "2) Envíos adicionales mientras escribiendo=true son bloqueados; no se agregan al historial ni al backend.\n"
         "3) Al responder el bot, escribiendo vuelve a false y el indicador desaparece.\n"
         "4) Solo un mensaje de usuario y una respuesta del bot en el historial por secuencia probada."),
        ("Salida obtenida",
         "El bloqueo 'if (this.escribiendo) return' funciona correctamente para el input de texto. Sin embargo, "
         "el método accionRapida() (línea 152-155 de chatbot.component.ts) invoca directamente enviarMensaje() "
         "sin verificar el estado escribiendo, produciendo en 2/10 ejecuciones un mensaje duplicado antes de que "
         "escribiendoSubject.next(false) sea emitido."),
        ("Resultado",                 "En seguimiento"),
        ("Seguimiento",               "BUG-BOT-004: Condición de carrera en accionRapida() de ChatbotComponent (línea 152-155). No verifica escribiendo$ antes de invocar enviarMensaje(). Se requiere agregar guard al inicio de accionRapida() o consolidar la lógica."),
        ("Severidad",                 "Medio"),
        ("Evidencia",
         "Grabación Cypress mostrando los dos mensajes duplicados en el historial. "
         "Screenshot de Network tab con dos peticiones HTTP duplicadas a /v1/chatbot/consulta con mismo payload. "
         "Screenshot de la UI con indicador 'El bot está escribiendo…' activo."),
        ("Firma de aprobación",       "Juan Pablo Giraldo Echavarria (Analista de QA)"),
    ],

    # ── CASO 005 ──────────────────────────────────────────────────────────────
    [
        ("Nombre del proyecto",       "SWO (ServiceDesk)"),
        ("Nombre del caso de prueba", "Carga masiva de 500 tickets históricos en Reportes — Rendimiento y estabilidad de UI"),
        ("Fecha de revisión",         "11/06/2026"),
        ("Descripción del caso",
         "Prueba de rendimiento y usabilidad bajo carga de datos. Se precargan 500 incidencias en PostgreSQL "
         "con Postman Collection Runner y se navega al módulo de Reportes (ReportsComponent). Se mide el tiempo "
         "de respuesta de /v1/estadisticas, la fluidez de las tarjetas métricas y la ejecución de descargarExcel() "
         "y descargarPDF() que solicitan GET /v1/incidencias?size=500. Fin: verificar que la UI no se bloquee "
         "durante la generación del PDF con jsPDF + jspdf-autotable."),
        ("Ambiente o entorno de prueba",
         "SO: Windows 11 Pro 22H2 | Navegador: Chrome 124 | Backend Spring Boot 3.x con HikariCP en Railway PostgreSQL | "
         "Dataset: 500 incidencias generadas con Postman Collection Runner | RAM disponible: 16 GB | Resolución: 1920x1080"),
        ("Herramienta utilizada",
         "Postman v11 (Collection Runner para precarga de 500 registros via POST /v1/incidencias) · "
         "Chrome DevTools — Performance tab (grabación 30 s) · "
         "Lighthouse (auditoría sobre /reportes) · "
         "Angular DevTools — detección de ciclos de change detection excesivos"),
        ("Autor del caso de prueba",  "Juan Pablo Giraldo Echavarria"),
        ("Número de caso",            "005"),
        ("Salida esperada",
         "1) Módulo carga estadísticas desde /v1/estadisticas en menos de 3 segundos.\n"
         "2) Tarjetas de métricas se renderizan sin bloquear el hilo principal del navegador.\n"
         "3) Exportación CSV se completa en menos de 5 s para 500 registros con BOM UTF-8 correcto.\n"
         "4) Generación de PDF con jsPDF + autoTable finaliza en menos de 8 s sin errores en consola.\n"
         "5) Toast muestra 'Excel generado: 500 incidencias' o 'PDF generado: 500 incidencias'."),
        ("Salida obtenida",
         "Estadísticas cargan en 1.8 s. Tarjetas métricas correctas. CSV se completa en 3.2 s con BOM correcto. "
         "Sin embargo, la generación del PDF provoca un bloqueo del hilo principal de ~4200 ms (Long Task en Performance tab), "
         "causando freeze visible de la UI. Después de la tarea larga, el PDF se descarga correctamente."),
        ("Resultado",                 "En seguimiento"),
        ("Seguimiento",               "BUG-REP-005: Generación sincrónica del PDF con jsPDF en descargarPDF() (reports.component.ts:156) bloquea el hilo principal ~4200ms con 500 registros. Recomendación: mover a Web Worker o implementar paginación de máx. 100 registros por lote."),
        ("Severidad",                 "Medio"),
        ("Evidencia",
         "Captura del Performance tab mostrando Long Task de 4200ms durante generación de PDF. "
         "Screenshot de Lighthouse con score de rendimiento de /reportes. "
         "PDF descargado correctamente con 500 incidencias en tabla autoTable. "
         "Grabación mostrando freeze visible de la UI durante la generación."),
        ("Firma de aprobación",       "Juan Pablo Giraldo Echavarria (Analista de QA)"),
    ],
]


# ─── Hoja de resumen ─────────────────────────────────────────────────────────

def write_summary(wb):
    ws = wb.create_sheet(title="Resumen Ejecutivo", index=0)
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 12

    # Título
    ws.merge_cells("A1:F1")
    c = ws.cell(row=1, column=1,
                value="RESUMEN EJECUTIVO DE CASOS DE PRUEBA — GA9-220501096-AA2-EV01")
    c.font = Font(bold=True, color=COLOR_RESUMEN_FG, size=13, name="Calibri")
    c.fill = PatternFill("solid", fgColor=COLOR_RESUMEN_HDR)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = THICK_BOTTOM
    ws.row_dimensions[1].height = 32

    ws.merge_cells("A2:F2")
    c = ws.cell(row=2, column=1,
                value="SWO (ServiceDesk)  ·  Analista: Juan Pablo Giraldo Echavarria  ·  Fecha: 11/06/2026")
    c.font = Font(bold=False, color="CCCCCC", size=9, name="Calibri")
    c.fill = PatternFill("solid", fgColor=COLOR_RESUMEN_HDR)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = BORDER
    ws.row_dimensions[2].height = 18

    # Cabecera tabla
    headers = ["N.°", "Módulo / Componente", "Tipo de Prueba",
               "Resultado", "Bug ID", "Severidad"]
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=3, column=col, value=h)
        c.font = Font(bold=True, color=COLOR_RESUMEN_FG, size=10, name="Calibri")
        c.fill = PatternFill("solid", fgColor=COLOR_RESUMEN_HDR)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = BORDER
    ws.row_dimensions[3].height = 22

    # Datos
    data = [
        ("001", "Incidentes — Móvil Android",    "Resiliencia Offline",       "En seguimiento", "BUG-MOV-001", "Medio"),
        ("002", "Incidentes — Web Angular 17",   "Validación Frontend",       "En seguimiento", "BUG-WEB-002", "Bajo"),
        ("003", "Usuarios / Proyectos — API",    "Seguridad (RBAC)",          "Rechazado",      "BUG-SEC-003", "Alto"),
        ("004", "Chatbot — Web",                 "Funcional / Concurrencia",  "En seguimiento", "BUG-BOT-004", "Medio"),
        ("005", "Reportes — Web",                "Rendimiento / UI",          "En seguimiento", "BUG-REP-005", "Medio"),
    ]

    result_colors = {
        "Rechazado":      ("F8D7DA", COLOR_RECHAZADO),
        "En seguimiento": ("FFF3CD", COLOR_SEGUIMIENTO),
        "Aprobado":       ("D4EDDA", COLOR_APROBADO),
    }
    sev_colors = {
        "Alto":  ("F8D7DA", COLOR_ALTO),
        "Medio": ("FFF3CD", COLOR_SEGUIMIENTO),
        "Bajo":  ("D4EDDA", COLOR_BAJO),
    }

    for i, row_data in enumerate(data):
        row = 4 + i
        ws.row_dimensions[row].height = 28
        alt = i % 2 == 0

        for col, val in enumerate(row_data, start=1):
            bg = COLOR_ALT_ROW if alt else "FFFFFF"
            fg = "0B1116"
            bold = False
            align_h = "left"

            if col == 4:  # Resultado
                bg, fg = result_colors.get(val, (bg, fg))
                bold = True
                align_h = "center"
            elif col == 6:  # Severidad
                bg, fg = sev_colors.get(val, (bg, fg))
                bold = True
                align_h = "center"
            elif col in (1, 5):
                align_h = "center"

            c = ws.cell(row=row, column=col, value=val)
            c.font = Font(bold=bold, color=fg, size=10, name="Calibri")
            c.fill = PatternFill("solid", fgColor=bg)
            c.alignment = Alignment(horizontal=align_h, vertical="center",
                                    wrap_text=True)
            c.border = BORDER

    # Nota al pie
    note_row = 10
    ws.merge_cells(f"A{note_row}:F{note_row}")
    c = ws.cell(row=note_row, column=1,
                value="NOTA: BUG-SEC-003 tiene severidad ALTA y requiere atención inmediata antes del despliegue a producción.")
    c.font = Font(bold=True, color=COLOR_RECHAZADO, size=9, name="Calibri",
                  italic=True)
    c.fill = PatternFill("solid", fgColor="F8D7DA")
    c.alignment = Alignment(horizontal="center", vertical="center",
                            wrap_text=True)
    c.border = BORDER
    ws.row_dimensions[note_row].height = 22


# ─── Main ─────────────────────────────────────────────────────────────────────

def main():
    wb = Workbook()
    # Eliminar hoja por defecto
    del wb["Sheet"]

    write_summary(wb)
    for idx, case_data in enumerate(CASOS, start=1):
        write_case(wb, idx, case_data)

    output_path = r"c:\Users\PruebasTrading\OneDrive\SENA\PROYECTOS\SWO\GA9-220501096-AA2-EV01_Casos_de_Prueba_SWO.xlsx"
    wb.save(output_path)
    print(f"Archivo generado: {output_path}")


if __name__ == "__main__":
    main()
